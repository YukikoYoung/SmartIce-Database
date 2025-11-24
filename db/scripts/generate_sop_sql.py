#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
野百灵餐饮集团 - SOP数据SQL生成器
从野百灵标准SOP.xlsx生成SQL录入脚本

数据来源:
  - 野百灵标准SOP.xlsx (7个sheet, 44个产品, 266条原料记录)

生成内容:
  1. SOP主表 (standard_operating_procedure)
  2. SOP原料明细表 (sop_ingredient)
  3. SOP操作步骤表 (sop_procedure)

注意: SOP表需要先在schema中创建
"""

from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# 路径配置
BASE_DIR = Path(__file__).parent
SOURCE_FILE = BASE_DIR / "野百灵标准SOP.xlsx"
OUTPUT_FILE = BASE_DIR / "data_sop_import.sql"

# SOP表结构DDL (如果不存在则创建)
SOP_DDL = """
-- ============================================================================
-- SOP相关表结构 (如果不存在则创建)
-- ============================================================================

-- SOP主表
CREATE TABLE IF NOT EXISTS standard_operating_procedure (
    sop_id SERIAL PRIMARY KEY,
    sop_code VARCHAR(50) UNIQUE NOT NULL,
    product_id INT,
    product_name VARCHAR(200) NOT NULL,
    category VARCHAR(100),
    version VARCHAR(20) DEFAULT 'v1.0',

    effective_date DATE DEFAULT CURRENT_DATE,
    expiry_date DATE,
    is_current BOOLEAN DEFAULT TRUE,

    preparation_time_minutes INT,
    difficulty_level VARCHAR(20),

    status VARCHAR(20) DEFAULT 'approved',
    created_by INT,
    approved_by INT,
    approved_at TIMESTAMP,

    notes TEXT,

    is_active BOOLEAN DEFAULT TRUE,
    created_at TIMESTAMP DEFAULT NOW(),
    updated_at TIMESTAMP,

    CONSTRAINT fk_sop_product FOREIGN KEY (product_id)
        REFERENCES product(product_id)
);

CREATE INDEX IF NOT EXISTS idx_sop_product ON standard_operating_procedure(product_id);
CREATE INDEX IF NOT EXISTS idx_sop_category ON standard_operating_procedure(category);
CREATE INDEX IF NOT EXISTS idx_sop_current ON standard_operating_procedure(is_current) WHERE is_current = TRUE;

-- SOP原料明细表
CREATE TABLE IF NOT EXISTS sop_ingredient (
    sop_ingredient_id SERIAL PRIMARY KEY,
    sop_id INT NOT NULL,
    ingredient_id INT,
    ingredient_name VARCHAR(200) NOT NULL,

    sequence_no INT DEFAULT 0,
    quantity DECIMAL(10,3),
    unit_id INT,
    unit_code VARCHAR(20),

    yield_rate DECIMAL(5,4) DEFAULT 1.0,

    preparation_note TEXT,
    operation_point TEXT,

    is_main_ingredient BOOLEAN DEFAULT FALSE,

    created_at TIMESTAMP DEFAULT NOW(),

    CONSTRAINT fk_sop_ing_sop FOREIGN KEY (sop_id)
        REFERENCES standard_operating_procedure(sop_id) ON DELETE CASCADE,
    CONSTRAINT fk_sop_ing_ingredient FOREIGN KEY (ingredient_id)
        REFERENCES product(product_id),
    CONSTRAINT fk_sop_ing_unit FOREIGN KEY (unit_id)
        REFERENCES unit_of_measure(unit_id)
);

CREATE INDEX IF NOT EXISTS idx_sop_ing_sop ON sop_ingredient(sop_id);
CREATE INDEX IF NOT EXISTS idx_sop_ing_ingredient ON sop_ingredient(ingredient_id);

-- SOP操作步骤表
CREATE TABLE IF NOT EXISTS sop_procedure (
    procedure_id SERIAL PRIMARY KEY,
    sop_id INT NOT NULL,

    step_no INT NOT NULL,
    step_name VARCHAR(200),
    step_description TEXT NOT NULL,

    duration_seconds INT,
    temperature_celsius DECIMAL(5,1),

    quality_standard TEXT,
    safety_note TEXT,

    image_url VARCHAR(500),
    video_url VARCHAR(500),

    created_at TIMESTAMP DEFAULT NOW(),

    CONSTRAINT fk_sop_proc_sop FOREIGN KEY (sop_id)
        REFERENCES standard_operating_procedure(sop_id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_sop_proc_sop ON sop_procedure(sop_id);

COMMENT ON TABLE standard_operating_procedure IS 'SOP主表 - 标准操作流程';
COMMENT ON TABLE sop_ingredient IS 'SOP原料明细表 - 每个SOP需要的原材料';
COMMENT ON TABLE sop_procedure IS 'SOP操作步骤表 - 每个SOP的具体操作步骤';
"""

def load_sop_data(wb):
    """加载SOP数据"""
    sop_list = []

    # 遍历所有sheet (每个sheet是一个品类)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_product = None

        for row in range(2, ws.max_row + 1):
            # 列: 序号, 产品名称, 食材名称, 单位, 使用量, 出成率, 操作要点
            seq_no = ws.cell(row=row, column=1).value
            product_name = ws.cell(row=row, column=2).value
            ingredient_name = ws.cell(row=row, column=3).value
            unit = ws.cell(row=row, column=4).value
            usage = ws.cell(row=row, column=5).value
            yield_rate = ws.cell(row=row, column=6).value
            operation_point = ws.cell(row=row, column=7).value

            # 新产品开始
            if product_name and str(product_name).strip():
                if current_product and current_product['ingredients']:
                    sop_list.append(current_product)

                current_product = {
                    'name': str(product_name).strip().replace('（', '(').replace('）', ')'),
                    'category': sheet_name,
                    'ingredients': []
                }

            # 添加原料
            if current_product and ingredient_name and str(ingredient_name).strip():
                ing_name = str(ingredient_name).strip()

                # 解析用量
                usage_qty = 0
                if usage:
                    usage_str = str(usage).strip()
                    if usage_str != '待补充':
                        try:
                            usage_qty = float(usage_str)
                        except:
                            pass

                # 解析单位
                unit_code = 'g'
                if unit:
                    unit_str = str(unit).strip().lower()
                    if unit_str == 'piece':
                        unit_code = 'piece'
                    elif unit_str == 'l':
                        unit_code = 'l'
                    elif unit_str == 'ml':
                        unit_code = 'ml'
                    elif unit_str == 'pack':
                        unit_code = 'pack'

                # 解析出成率
                yr = 1.0
                if yield_rate:
                    try:
                        yr_val = float(yield_rate)
                        yr = yr_val if yr_val <= 1 else yr_val / 100
                    except:
                        yr = 1.0

                current_product['ingredients'].append({
                    'name': ing_name,
                    'quantity': usage_qty,
                    'unit': unit_code,
                    'yield_rate': yr,
                    'operation_point': str(operation_point).strip() if operation_point else ''
                })

        # 添加最后一个产品
        if current_product and current_product['ingredients']:
            sop_list.append(current_product)

    return sop_list

def escape_sql(s):
    """转义SQL字符串"""
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"

def generate_sql():
    """生成SQL脚本"""
    print(f"📖 读取SOP文件: {SOURCE_FILE}")
    wb = load_workbook(SOURCE_FILE, data_only=True)

    sop_list = load_sop_data(wb)
    total_ingredients = sum(len(s['ingredients']) for s in sop_list)

    print(f"📊 SOP数量: {len(sop_list)}")
    print(f"📊 原料记录总数: {total_ingredients}")

    wb.close()

    sql_lines = []

    # 文件头
    sql_lines.append(f"""-- ============================================================================
-- 野百灵餐饮集团 - SOP数据导入
-- ============================================================================
-- 版本: v1.0.0
-- 数据来源: 野百灵标准SOP.xlsx
-- SOP数量: {len(sop_list)}
-- 原料记录总数: {total_ingredients}
-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
-- ============================================================================

-- 设置加密密钥
SET app.encryption_key = 'ybl-restaurant-encryption-key-2025';
""")

    # 添加表结构DDL
    sql_lines.append(SOP_DDL)

    # SOP主表数据
    sql_lines.append("""
-- ============================================================================
-- 1. SOP主表数据
-- ============================================================================
""")

    for i, sop in enumerate(sop_list, 1):
        sop_code = f"SOP-{i:03d}"

        sql_lines.append(f"""
-- {i}. {sop['name']} (品类:{sop['category']}, 原料:{len(sop['ingredients'])}种)
INSERT INTO standard_operating_procedure (
    sop_code, product_name, category, version,
    effective_date, is_current, status, is_active
)
SELECT
    '{sop_code}',
    {escape_sql(sop['name'])},
    {escape_sql(sop['category'])},
    'v1.0',
    '2025-01-01',
    TRUE,
    'approved',
    TRUE
WHERE NOT EXISTS (
    SELECT 1 FROM standard_operating_procedure WHERE sop_code = '{sop_code}'
);

-- 关联产品ID (如果存在)
UPDATE standard_operating_procedure
SET product_id = (
    SELECT product_id FROM product
    WHERE product_name = {escape_sql(sop['name'])}
      AND product_type = 'finished'
    LIMIT 1
)
WHERE sop_code = '{sop_code}'
  AND product_id IS NULL;
""")

    # SOP原料明细
    sql_lines.append("""
-- ============================================================================
-- 2. SOP原料明细
-- ============================================================================
""")

    for i, sop in enumerate(sop_list, 1):
        sop_code = f"SOP-{i:03d}"

        sql_lines.append(f"""
-- SOP原料: {sop['name']}
DO $$
DECLARE
    v_sop_id INT;
    v_ingredient_id INT;
BEGIN
    SELECT sop_id INTO v_sop_id FROM standard_operating_procedure WHERE sop_code = '{sop_code}';

    IF v_sop_id IS NULL THEN
        RAISE NOTICE 'SOP不存在: {sop_code}';
        RETURN;
    END IF;

    -- 清空原有原料明细
    DELETE FROM sop_ingredient WHERE sop_id = v_sop_id;
""")

        for j, ing in enumerate(sop['ingredients'], 1):
            yield_info = f", 出成率:{ing['yield_rate']*100:.0f}%" if ing['yield_rate'] < 1.0 else ""
            qty_info = ing['quantity'] if ing['quantity'] > 0 else '待补充'

            sql_lines.append(f"""
    -- 原料 {j}: {ing['name']} ({qty_info}{ing['unit']}{yield_info})
    SELECT product_id INTO v_ingredient_id
    FROM product
    WHERE product_name = {escape_sql(ing['name'])}
      AND product_type = 'raw_material'
    LIMIT 1;

    INSERT INTO sop_ingredient (
        sop_id, ingredient_id, ingredient_name, sequence_no,
        quantity, unit_code, yield_rate, operation_point, is_main_ingredient
    ) VALUES (
        v_sop_id,
        v_ingredient_id,
        {escape_sql(ing['name'])},
        {j},
        {ing['quantity'] if ing['quantity'] > 0 else 'NULL'},
        {escape_sql(ing['unit'])},
        {ing['yield_rate']},
        {escape_sql(ing['operation_point']) if ing['operation_point'] else 'NULL'},
        {'TRUE' if j == 1 else 'FALSE'}
    );
""")

        sql_lines.append("""
END $$;
""")

    # 统计信息
    sql_lines.append(f"""
-- ============================================================================
-- 3. 数据验证
-- ============================================================================
-- 验证SOP数量
-- SELECT COUNT(*) AS sop_count FROM standard_operating_procedure WHERE is_current = TRUE;

-- 验证原料关联率
-- SELECT
--     COUNT(*) AS total_ingredients,
--     COUNT(ingredient_id) AS linked_ingredients,
--     ROUND(COUNT(ingredient_id)::NUMERIC / COUNT(*) * 100, 2) AS link_rate
-- FROM sop_ingredient;

-- 查看缺失用量的原料
-- SELECT si.ingredient_name, sop.product_name, sop.category
-- FROM sop_ingredient si
-- JOIN standard_operating_procedure sop ON si.sop_id = sop.sop_id
-- WHERE si.quantity IS NULL OR si.quantity = 0
-- ORDER BY sop.category, sop.product_name;

-- ============================================================================
-- 脚本完成
-- ============================================================================
""")

    # 写入文件
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_lines))

    print(f"✅ SQL脚本已生成: {OUTPUT_FILE}")

    # 统计缺失用量的数量
    missing_qty = sum(1 for s in sop_list for i in s['ingredients'] if i['quantity'] == 0)
    print(f"⚠️  缺失用量的原料: {missing_qty} 条")

if __name__ == '__main__':
    generate_sql()
