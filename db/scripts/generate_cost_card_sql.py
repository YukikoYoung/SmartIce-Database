#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
野百灵餐饮集团 - 成本卡数据SQL生成器
从野百灵产品标准成本卡.xlsx生成SQL录入脚本

数据来源:
  - 原材料数据: 原材料价格表 sheet (241条)
  - 产品配方数据: 肉类、海鲜（涮菜）sheet + 糖水铺 sheet (73个产品)

功能:
  1. 更新/新增原材料产品数据 (product表)
  2. 更新出成率数据 (std_yield_rate字段)
  3. 生成产品配方数据 (recipe + recipe_item表)
"""

import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# 路径配置
BASE_DIR = Path(__file__).parent
SOURCE_FILE = BASE_DIR / "野百灵产品标准成本卡.xlsx"
OUTPUT_FILE = BASE_DIR / "data_cost_card_import.sql"

def load_raw_materials(wb):
    """加载原材料价格表数据"""
    ws = wb['原材料价格表']
    materials = []

    for row in range(2, ws.max_row + 1):
        category = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        supplier = ws.cell(row=row, column=3).value
        brand = ws.cell(row=row, column=4).value
        spec = ws.cell(row=row, column=5).value
        base_qty = ws.cell(row=row, column=6).value
        purchase_price = ws.cell(row=row, column=7).value
        storage = ws.cell(row=row, column=8).value
        yield_rate = ws.cell(row=row, column=9).value
        base_price = ws.cell(row=row, column=10).value

        if name and str(name).strip():
            # 解析规格中的单位
            base_unit = 'g'  # 默认克
            if spec:
                spec_str = str(spec)
                if 'ml' in spec_str.lower():
                    base_unit = 'ml'
                elif '个' in spec_str or '只' in spec_str or '颗' in spec_str:
                    base_unit = 'piece'

            # 安全转换数值
            def safe_float(val, default=0):
                if val is None:
                    return default
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return default

            materials.append({
                'category': str(category).strip() if category else '其他',
                'name': str(name).strip(),
                'supplier': str(supplier).strip() if supplier else '-',
                'brand': str(brand).strip() if brand else '-',
                'spec': str(spec).strip() if spec else '-',
                'base_qty': safe_float(base_qty, 500),
                'purchase_price': safe_float(purchase_price, 0),
                'storage': str(storage).strip() if storage else '常温',
                'yield_rate': safe_float(yield_rate, 1.0),
                'base_price': safe_float(base_price, 0),
                'base_unit': base_unit
            })

    return materials

def load_products(wb, sheet_name):
    """加载产品配方数据

    成本卡格式:
    - 第1列: 产品名称 (只在产品首行有值)
    - 第2列: 原材料名称
    - 后续列: 根据sheet不同结构不同
    - "售价"行表示产品配方结束
    """
    import re
    ws = wb[sheet_name]
    products = []
    current_product = None

    for row in range(3, ws.max_row + 1):  # 从第3行开始（跳过标题行）
        # 第1列: 产品名称
        product_name = ws.cell(row=row, column=1).value
        # 第2列: 原材料名称
        ingredient_name = ws.cell(row=row, column=2).value

        # 检查是否是"售价"行（表示产品结束）
        if ingredient_name and str(ingredient_name).strip() == '售价':
            continue

        # 新产品开始（第1列有产品名称）
        if product_name and str(product_name).strip():
            # 保存上一个产品
            if current_product and current_product['ingredients']:
                products.append(current_product)

            current_product = {
                'name': str(product_name).strip().replace('（', '(').replace('）', ')'),
                'sheet': sheet_name,
                'ingredients': []
            }

        # 添加原材料
        if current_product and ingredient_name and str(ingredient_name).strip():
            ing_name = str(ingredient_name).strip()

            # 跳过"售价"、"总成本"等非原材料行
            if ing_name in ['售价', '总成本', '售卖份量/g', '总成本（元）']:
                continue

            # 默认值
            usage_qty = 0
            usage_unit = 'g'
            yr = 1.0

            # 糖水铺的出成率在第7列
            if sheet_name == '糖水铺':
                yield_rate = ws.cell(row=row, column=7).value
                if yield_rate:
                    try:
                        yr_val = float(yield_rate)
                        yr = yr_val if yr_val <= 1 else yr_val / 100
                    except:
                        yr = 1.0

            current_product['ingredients'].append({
                'name': ing_name,
                'quantity': usage_qty,
                'unit': usage_unit,
                'yield_rate': yr
            })

    # 添加最后一个产品
    if current_product and current_product['ingredients']:
        products.append(current_product)

    return products

def escape_sql(s):
    """转义SQL字符串"""
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"

def get_category_code(category):
    """获取品类编码"""
    category_map = {
        '肉类': 'CAT-MEAT',
        '蔬菜': 'CAT-VEG',
        '干杂': 'CAT-DRY',
        '海鲜': 'CAT-SEAFOOD',
        '酱料': 'CAT-SAUCE',
        '酒水': 'CAT-LIQUOR',
        '糖水铺': 'CAT-DESSERT',
        '自制': 'CAT-SELFMADE',
        '集采': 'CAT-PURCHASE',
        '其他': 'CAT-OTHER'
    }
    return category_map.get(category, 'CAT-OTHER')

def generate_sql():
    """生成SQL脚本"""
    print(f"📖 读取成本卡文件: {SOURCE_FILE}")
    wb = load_workbook(SOURCE_FILE, data_only=True)

    # 加载数据
    materials = load_raw_materials(wb)
    products_meat = load_products(wb, '肉类、海鲜（涮菜）')
    products_dessert = load_products(wb, '糖水铺')
    all_products = products_meat + products_dessert

    print(f"📊 原材料数量: {len(materials)}")
    print(f"📊 产品数量: {len(all_products)} (肉类海鲜:{len(products_meat)}, 糖水铺:{len(products_dessert)})")

    wb.close()

    sql_lines = []

    # 文件头
    sql_lines.append(f"""-- ============================================================================
-- 野百灵餐饮集团 - 成本卡数据导入
-- ============================================================================
-- 版本: v1.0.0
-- 数据来源: 野百灵产品标准成本卡.xlsx
-- 原材料数量: {len(materials)}
-- 产品数量: {len(all_products)}
-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
-- ============================================================================

-- 设置加密密钥
SET app.encryption_key = 'ybl-restaurant-encryption-key-2025';

-- ============================================================================
-- 1. 原材料出成率更新
-- ============================================================================
-- 从成本卡更新原材料的标准出成率(std_yield_rate)
-- ============================================================================
""")

    # 更新原材料出成率
    updated_count = 0
    for mat in materials:
        if mat['yield_rate'] < 1.0:
            sql_lines.append(f"""
UPDATE product SET std_yield_rate = {mat['yield_rate']}
WHERE product_name = {escape_sql(mat['name'])}
  AND product_type = 'raw_material'
  AND std_yield_rate IS DISTINCT FROM {mat['yield_rate']};""")
            updated_count += 1

    sql_lines.append(f"""
-- 出成率<100%的原材料: {updated_count} 种

-- ============================================================================
-- 2. 新增原材料 (成本卡中有但数据库中没有的)
-- ============================================================================
""")

    # 新增原材料
    for i, mat in enumerate(materials, 1):
        product_code = f"RM-CC-{i:03d}"
        price_per_unit = mat['base_price'] / mat['base_qty'] if mat['base_qty'] > 0 else 0

        sql_lines.append(f"""
-- {i}. {mat['name']} (品类:{mat['category']}, 出成率:{mat['yield_rate']*100:.0f}%)
INSERT INTO product (
    product_code, product_name, product_type, category_id,
    is_saleable, is_ingredient, base_unit_id,
    current_cost_encrypted, std_yield_rate, storage_condition,
    description, is_active
)
SELECT
    '{product_code}',
    {escape_sql(mat['name'])},
    'raw_material',
    (SELECT category_id FROM product_category WHERE category_code = '{get_category_code(mat['category'])}'),
    FALSE,
    TRUE,
    (SELECT unit_id FROM unit_of_measure WHERE unit_code = '{mat['base_unit']}'),
    encrypt_cost({price_per_unit:.6f}),
    {mat['yield_rate']},
    {escape_sql(mat['storage'])},
    '来源:{mat['supplier']} | 规格:{mat['spec']} | 采购价:{mat['purchase_price']}元',
    TRUE
WHERE NOT EXISTS (
    SELECT 1 FROM product WHERE product_name = {escape_sql(mat['name'])} AND product_type = 'raw_material'
);""")

    # 产品配方
    sql_lines.append(f"""

-- ============================================================================
-- 3. 成品产品数据 (来自成本卡)
-- ============================================================================
""")

    for i, product in enumerate(all_products, 1):
        product_code = f"FP-CC-{i:03d}"

        # 计算产品总成本 (暂时设为0，需要根据配方计算)
        sql_lines.append(f"""
-- {i}. {product['name']} (原材料:{len(product['ingredients'])}种, 来源:{product['sheet']})
INSERT INTO product (
    product_code, product_name, product_type, category_id,
    is_saleable, is_ingredient, base_unit_id,
    cost_calculation_method, description, is_active
)
SELECT
    '{product_code}',
    {escape_sql(product['name'])},
    'finished',
    (SELECT category_id FROM product_category WHERE category_code = 'CAT-FINISHED'),
    TRUE,
    FALSE,
    (SELECT unit_id FROM unit_of_measure WHERE unit_code = 'serving'),
    'standard',
    '来源: {product['sheet']} | 原材料种类: {len(product["ingredients"])}种',
    TRUE
WHERE NOT EXISTS (
    SELECT 1 FROM product WHERE product_name = {escape_sql(product['name'])} AND product_type = 'finished'
)
ON CONFLICT (product_code) DO UPDATE SET
    product_name = EXCLUDED.product_name,
    description = EXCLUDED.description,
    updated_at = NOW();""")

    # 配方数据
    sql_lines.append(f"""

-- ============================================================================
-- 4. 配方主表
-- ============================================================================
""")

    for i, product in enumerate(all_products, 1):
        product_code = f"FP-CC-{i:03d}"
        recipe_code = f"RCP-CC-{i:03d}-v1"

        sql_lines.append(f"""
-- 配方: {product['name']}
INSERT INTO recipe (
    recipe_code, product_id, version, recipe_name,
    effective_date, is_current, yield_quantity, yield_unit_id, status
)
SELECT
    '{recipe_code}',
    p.product_id,
    'v1.0',
    {escape_sql(product['name'] + ' 标准配方(成本卡)')},
    '2025-01-01',
    TRUE,
    1.0,
    (SELECT unit_id FROM unit_of_measure WHERE unit_code = 'serving'),
    'approved'
FROM product p
WHERE p.product_name = {escape_sql(product['name'])}
  AND p.product_type = 'finished'
ON CONFLICT (recipe_code) DO UPDATE SET
    recipe_name = EXCLUDED.recipe_name,
    updated_at = NOW();""")

    # 配方明细
    sql_lines.append(f"""

-- ============================================================================
-- 5. 配方明细 (recipe_item)
-- ============================================================================
""")

    for i, product in enumerate(all_products, 1):
        recipe_code = f"RCP-CC-{i:03d}-v1"

        sql_lines.append(f"""
-- 配方明细: {product['name']}
DO $$
DECLARE
    v_recipe_id INT;
    v_ingredient_id INT;
BEGIN
    SELECT recipe_id INTO v_recipe_id FROM recipe WHERE recipe_code = '{recipe_code}';

    IF v_recipe_id IS NULL THEN
        RAISE NOTICE '配方不存在: {recipe_code}';
        RETURN;
    END IF;

    -- 清空原有配方明细
    DELETE FROM recipe_item WHERE recipe_id = v_recipe_id;
""")

        for j, ing in enumerate(product['ingredients'], 1):
            sql_lines.append(f"""
    -- 原材料 {j}: {ing['name']} ({ing['quantity']}{ing['unit']}, 出成率:{ing['yield_rate']*100:.0f}%)
    SELECT product_id INTO v_ingredient_id
    FROM product
    WHERE product_name = {escape_sql(ing['name'])}
      AND product_type = 'raw_material'
    LIMIT 1;

    IF v_ingredient_id IS NOT NULL THEN
        INSERT INTO recipe_item (
            recipe_id, ingredient_id, ingredient_type, usage_sequence,
            quantity, unit_id, usage_yield_rate, is_main_ingredient
        ) VALUES (
            v_recipe_id,
            v_ingredient_id,
            'raw_material',
            {j},
            {ing['quantity']},
            (SELECT unit_id FROM unit_of_measure WHERE unit_code = '{ing['unit']}'),
            {ing['yield_rate']},
            {'TRUE' if j == 1 else 'FALSE'}
        );
    ELSE
        RAISE NOTICE '原材料未找到: {ing["name"]}';
    END IF;
""")

        sql_lines.append("""
END $$;
""")

    # 验证脚本
    sql_lines.append("""
-- ============================================================================
-- 6. 数据验证
-- ============================================================================
-- 验证导入的原材料数量
-- SELECT COUNT(*) AS raw_material_count FROM product WHERE product_type = 'raw_material';

-- 验证导入的成品数量
-- SELECT COUNT(*) AS finished_product_count FROM product WHERE product_type = 'finished';

-- 验证配方完整性
-- SELECT
--     r.recipe_code,
--     p.product_name,
--     COUNT(ri.recipe_item_id) AS ingredient_count
-- FROM recipe r
-- JOIN product p ON r.product_id = p.product_id
-- LEFT JOIN recipe_item ri ON r.recipe_id = ri.recipe_id
-- WHERE r.is_current = TRUE
-- GROUP BY r.recipe_code, p.product_name
-- ORDER BY r.recipe_code;

-- ============================================================================
-- 脚本完成
-- ============================================================================
""")

    # 写入文件
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_lines))

    print(f"✅ SQL脚本已生成: {OUTPUT_FILE}")
    print(f"📊 原材料出成率更新: {updated_count} 种")
    print(f"📊 产品配方: {len(all_products)} 个")

if __name__ == '__main__':
    generate_sql()
