#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
野百灵餐饮集团 - 成品菜与配方数据SQL生成器
从成本卡JSON生成SQL录入脚本

数据来源:
  - 配方数据: 成本卡_产品配方_最终修正版.json
  - 原材料价格: 原材料价格及规格明细表.md (标准价格源)
  - 出成率数据: WL成本卡.xlsx (原材料价格表 sheet)

成本计算公式:
  实际成本 = 用量(g) × 单价(元/g) / 出成率

  示例: 吊龙 250g, 单价0.096元/g, 出成率90%
  实际成本 = 250 × 0.096 / 0.9 = 26.67元
"""

import json
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# 路径配置
BASE_DIR = Path(__file__).parent
SOURCE_FILE = BASE_DIR.parent / "数据排查与审计专项" / "成本卡_产品配方_最终修正版.json"
PRICE_FILE = BASE_DIR / "原材料价格及规格明细表.md"
YIELD_FILE = BASE_DIR / "WL成本卡.xlsx"
OUTPUT_FILE = BASE_DIR / "data_products_recipes.sql"

def load_price_standard():
    """加载原材料价格标准表"""
    price_dict = {}

    with open(PRICE_FILE, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')
    for line in lines:
        if line.startswith('|') and not line.startswith('|---') and '原材料名称' not in line:
            parts = [p.strip() for p in line.split('|')]
            if len(parts) >= 9:
                name = parts[1]
                base_unit = parts[5]  # 基础单位
                base_qty = parts[6]   # 基础单位数量
                base_price = parts[7] # 基础单价
                price_per_unit = parts[8]  # 单价/克或单价/ml或单价/个

                try:
                    price_dict[name] = {
                        'base_unit': base_unit,
                        'base_qty': float(base_qty) if base_qty else 0,
                        'base_price': float(base_price) if base_price else 0,
                        'price_per_unit': float(price_per_unit) if price_per_unit else 0
                    }
                except ValueError:
                    pass

    return price_dict

def load_yield_data():
    """加载出成率数据从WL成本卡.xlsx"""
    yield_dict = {}

    try:
        wb = load_workbook(YIELD_FILE, data_only=True)
        ws = wb['原材料价格表']

        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=2).value  # 原材料名称
            yield_rate = ws.cell(row=row, column=9).value  # 标准出成率

            if name and yield_rate:
                yield_dict[str(name).strip()] = float(yield_rate)

        wb.close()
    except Exception as e:
        print(f"⚠️ 无法加载出成率数据: {e}")

    return yield_dict

def load_recipe_data():
    """加载配方数据"""
    with open(SOURCE_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def normalize_name(name):
    """规范化名称：全角转半角"""
    return name.replace('（', '(').replace('）', ')').strip()

def escape_sql(s):
    """转义SQL字符串"""
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"

def match_ingredient_name(name, price_dict):
    """尝试匹配原材料名称"""
    # 直接匹配
    if name in price_dict:
        return name, price_dict[name]

    # 规范化后匹配
    normalized = normalize_name(name)
    if normalized in price_dict:
        return normalized, price_dict[normalized]

    # 别名映射
    alias_map = {
        '10头鲍': '8头鲍',
        '耗油': '蚝油',
        '啤酒': '雪花纯生',
        '玫瑰乳扇': '乳扇',
        '甜玉糊': '甜玉米',
        '野菜卷（成品）': '野菜卷',
        '三黄鸡': '鸡肉',
        '酸辣酱': '蒜蓉辣椒酱',  # 近似
        '老凯里非遗酸汤': '酸汤底料',
        '黄辣丁': '罗非鱼',  # 按鱼类计算
    }

    if name in alias_map:
        mapped_name = alias_map[name]
        if mapped_name in price_dict:
            return mapped_name, price_dict[mapped_name]

    return None, None

def get_yield_rate(name, yield_dict, price_dict):
    """获取原材料的出成率"""
    # 直接匹配
    if name in yield_dict:
        return yield_dict[name]

    # 规范化后匹配
    normalized = normalize_name(name)
    if normalized in yield_dict:
        return yield_dict[normalized]

    # 别名映射后匹配
    alias_map = {
        '10头鲍': '8头鲍',
        '耗油': '蚝油',
        '啤酒': '雪花纯生',
        '玫瑰乳扇': '乳扇',
        '甜玉糊': '甜玉米',
        '野菜卷（成品）': '野菜卷',
        '三黄鸡': '鸡肉',
        '酸辣酱': '蒜蓉辣椒酱',
        '老凯里非遗酸汤': '酸汤底料',
        '黄辣丁': '罗非鱼',
    }
    if name in alias_map and alias_map[name] in yield_dict:
        return yield_dict[alias_map[name]]

    # 默认100%出成率
    return 1.0

def generate_sql():
    """生成SQL脚本"""
    price_dict = load_price_standard()
    yield_dict = load_yield_data()
    data = load_recipe_data()
    products = data['产品数据']

    print(f"📊 原材料价格标准表: {len(price_dict)} 种")
    print(f"📊 出成率数据: {len(yield_dict)} 种 (其中<100%: {sum(1 for v in yield_dict.values() if v < 1)} 种)")

    sql_lines = []
    unmatched_ingredients = set()

    # 统计出成率<100%的原材料数量
    yield_count = sum(1 for v in yield_dict.values() if v < 1)

    # 文件头
    sql_lines.append(f"""-- ============================================================================
-- 野百灵餐饮集团 - 成品菜与配方数据录入
-- ============================================================================
-- 版本: v1.2.0
-- 数据来源:
--   - 配方: 成本卡_产品配方_最终修正版.json
--   - 价格: 原材料价格及规格明细表.md (标准价格源)
--   - 出成率: WL成本卡.xlsx (原材料价格表)
-- 产品数量: {len(products)}
-- 出成率<100%原材料: {yield_count}种
-- 成本公式: 实际成本 = 用量 × 单价 / 出成率
-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
-- ============================================================================

-- 设置加密密钥
SET app.encryption_key = 'ybl-restaurant-encryption-key-2025';

-- ============================================================================
-- 1. 成品菜产品数据 (product_type = 'finished')
-- ============================================================================
""")

    # 先计算每个产品的标准成本 (考虑出成率)
    product_costs = {}
    for i, product in enumerate(products, 1):
        product_name = normalize_name(product['产品名称'])
        total_cost = 0

        for ingredient in product['原材料']:
            ing_name = ingredient['名称']
            ing_qty = ingredient['用量_g']

            matched_name, price_info = match_ingredient_name(ing_name, price_dict)
            if price_info:
                # 使用标准价格计算成本
                unit_price = price_info['price_per_unit']
                # 获取出成率并计算实际成本: 成本 = 用量 × 单价 / 出成率
                yield_rate = get_yield_rate(ing_name, yield_dict, price_dict)
                ing_cost = ing_qty * unit_price / yield_rate
            else:
                # 使用原配方成本
                ing_cost = ingredient['成本_元']
                unmatched_ingredients.add(ing_name)

            total_cost += ing_cost

        product_costs[product_name] = round(total_cost, 2)

    # 生成产品INSERT
    for i, product in enumerate(products, 1):
        product_code = f"FP-{i:03d}"
        product_name = normalize_name(product['产品名称'])
        total_cost = product_costs[product_name]
        ingredient_count = product['原材料种类']

        sql_lines.append(f"""
-- {i}. {product_name} (标准成本: {total_cost}元)
INSERT INTO product (
    product_code, product_name, product_type, category_id,
    is_saleable, is_ingredient, base_unit_id,
    current_cost_encrypted, cost_calculation_method,
    description, is_active
) VALUES (
    '{product_code}',
    {escape_sql(product_name)},
    'finished',
    (SELECT category_id FROM product_category WHERE category_code = 'CAT-FINISHED'),
    TRUE,
    FALSE,
    (SELECT unit_id FROM unit_of_measure WHERE unit_code = 'serving'),
    encrypt_cost({total_cost}),
    'standard',
    '原材料种类: {ingredient_count}种 | 基于标准价格表计算',
    TRUE
) ON CONFLICT (product_code) DO UPDATE SET
    product_name = EXCLUDED.product_name,
    current_cost_encrypted = EXCLUDED.current_cost_encrypted,
    description = EXCLUDED.description,
    updated_at = NOW();
""")

    # 配方表头
    sql_lines.append("""
-- ============================================================================
-- 2. 配方主表 (recipe)
-- ============================================================================
""")

    # 生成配方INSERT
    for i, product in enumerate(products, 1):
        product_code = f"FP-{i:03d}"
        recipe_code = f"RCP-{i:03d}-v1"
        product_name = normalize_name(product['产品名称'])
        total_cost = product_costs[product_name]

        sql_lines.append(f"""
-- 配方: {product_name}
INSERT INTO recipe (
    recipe_code, product_id, version, recipe_name,
    effective_date, is_current,
    yield_quantity, yield_unit_id,
    total_material_cost_encrypted, total_cost_encrypted,
    status
) VALUES (
    '{recipe_code}',
    (SELECT product_id FROM product WHERE product_code = '{product_code}'),
    'v1.0',
    {escape_sql(product_name + ' 标准配方')},
    '2025-01-01',
    TRUE,
    1.0,
    (SELECT unit_id FROM unit_of_measure WHERE unit_code = 'serving'),
    encrypt_cost({total_cost}),
    encrypt_cost({total_cost}),
    'approved'
) ON CONFLICT (recipe_code) DO UPDATE SET
    total_material_cost_encrypted = EXCLUDED.total_material_cost_encrypted,
    total_cost_encrypted = EXCLUDED.total_cost_encrypted,
    updated_at = NOW();
""")

    # 配方明细表头
    sql_lines.append("""
-- ============================================================================
-- 3. 配方明细 (recipe_item)
-- ============================================================================
-- 价格来源: 原材料价格及规格明细表.md
-- 出成率来源: WL成本卡.xlsx
-- 成本公式: 实际成本 = 用量 × 单价 / 出成率
-- ============================================================================
""")

    # 生成配方明细INSERT
    for i, product in enumerate(products, 1):
        recipe_code = f"RCP-{i:03d}-v1"
        product_name = normalize_name(product['产品名称'])

        sql_lines.append(f"""
-- 配方明细: {product_name}
DO $$
DECLARE
    v_recipe_id INT;
    v_ingredient_id INT;
BEGIN
    SELECT recipe_id INTO v_recipe_id FROM recipe WHERE recipe_code = '{recipe_code}';

    IF v_recipe_id IS NULL THEN
        RAISE NOTICE '配方不存在: {recipe_code}';
        RETURN;
    END IF;

    DELETE FROM recipe_item WHERE recipe_id = v_recipe_id;
""")

        for j, ingredient in enumerate(product['原材料'], 1):
            ing_name = ingredient['名称']
            ing_qty = ingredient['用量_g']

            # 使用标准价格
            matched_name, price_info = match_ingredient_name(ing_name, price_dict)
            if price_info:
                unit_price = price_info['price_per_unit']
                db_name = matched_name  # 使用匹配后的名称查询数据库
            else:
                # 回退到原配方成本
                unit_price = ingredient['成本_元'] / ing_qty if ing_qty > 0 else 0
                db_name = ing_name

            # 获取出成率并计算实际成本: 成本 = 用量 × 单价 / 出成率
            yield_rate = get_yield_rate(ing_name, yield_dict, price_dict)
            ing_cost = ing_qty * unit_price / yield_rate

            # 生成出成率显示（100%时省略）
            yield_info = f", 出成率:{yield_rate*100:.0f}%" if yield_rate < 1.0 else ""

            sql_lines.append(f"""
    -- 原材料 {j}: {ing_name} (用量:{ing_qty}g, 单价:{unit_price:.4f}/g{yield_info}, 成本:{ing_cost:.2f}元)
    SELECT product_id INTO v_ingredient_id
    FROM product
    WHERE product_name = {escape_sql(db_name)}
      AND product_type = 'raw_material'
    LIMIT 1;

    IF v_ingredient_id IS NOT NULL THEN
        INSERT INTO recipe_item (
            recipe_id, ingredient_id, ingredient_type, usage_sequence,
            quantity, unit_id, unit_price_encrypted,
            usage_yield_rate, waste_rate, subtotal_cost_encrypted,
            is_main_ingredient
        ) VALUES (
            v_recipe_id,
            v_ingredient_id,
            'raw_material',
            {j},
            {ing_qty},
            (SELECT unit_id FROM unit_of_measure WHERE unit_code = 'g'),
            encrypt_cost({unit_price:.6f}),
            {yield_rate},
            0,
            encrypt_cost({ing_cost:.4f}),
            {'TRUE' if j == 1 else 'FALSE'}
        );
    ELSE
        RAISE NOTICE '原材料未找到: {db_name}';
    END IF;
""")

        sql_lines.append("""
END $$;
""")

    # 验证脚本
    sql_lines.append("""
-- ============================================================================
-- 4. 数据验证
-- ============================================================================
-- SELECT
--     p.product_code,
--     p.product_name,
--     r.recipe_code,
--     COUNT(ri.recipe_item_id) AS ingredient_count,
--     ROUND(SUM(decrypt_cost(ri.subtotal_cost_encrypted))::NUMERIC, 2) AS calculated_cost,
--     ROUND(decrypt_cost(p.current_cost_encrypted)::NUMERIC, 2) AS product_cost
-- FROM product p
-- LEFT JOIN recipe r ON p.product_id = r.product_id AND r.is_current = TRUE
-- LEFT JOIN recipe_item ri ON r.recipe_id = ri.recipe_id
-- WHERE p.product_type = 'finished'
-- GROUP BY p.product_id, p.product_code, p.product_name, r.recipe_code, p.current_cost_encrypted
-- ORDER BY p.product_code;

-- ============================================================================
-- 脚本完成
-- ============================================================================
""")

    # 写入文件
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_lines))

    print(f"✅ SQL脚本已生成: {OUTPUT_FILE}")
    print(f"📊 产品数量: {len(products)}")

    total_ingredients = sum(len(p['原材料']) for p in products)
    print(f"📊 配方明细总数: {total_ingredients}")

    if unmatched_ingredients:
        print(f"\n⚠️  未匹配的原材料 ({len(unmatched_ingredients)}种):")
        for name in sorted(unmatched_ingredients):
            print(f"   - {name}")
        print("   (这些原材料使用原配方成本)")

if __name__ == '__main__':
    generate_sql()
